#!/usr/bin/env python3

import json
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


DESKTOP = Path("/Users/nicolas/Desktop")
RESERVAS_XLSX = DESKTOP / "reservas_453521_1773150124.xlsx"
CLIENTES_XLSX = DESKTOP / "BASE DE DATOS CLIENTES AL 10-03.xlsx"
OUTPUT_SQL = DESKTOP / "importar_reservas_siluetachic.sql"
OUTPUT_JSON = DESKTOP / "importar_reservas_siluetachic_preview.json"


def sql_literal(value):
    if value is None:
        return "NULL"
    text = str(value).replace("\\", "\\\\").replace("'", "''")
    return f"'{text}'"


def normalize_text(value):
    if value is None:
        return None
    text = " ".join(str(value).split()).strip()
    return text or None


def normalize_email(value):
    text = normalize_text(value)
    return text.lower() if text else None


def normalize_phone(value):
    return normalize_text(value)


def normalize_rut(value):
    text = normalize_text(value)
    if not text:
        return None

    cleaned = "".join(ch for ch in text if ch.isdigit() or ch.lower() == "k").upper()
    if len(cleaned) < 2:
        return text

    body = cleaned[:-1]
    dv = cleaned[-1]
    parts = []
    while body:
        parts.append(body[-3:])
        body = body[:-3]
    return ".".join(reversed(parts)) + "-" + dv


def normalize_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = normalize_text(value)
    if text and len(text) >= 10:
        return text[:10]
    return None


def normalize_time(value):
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    text = normalize_text(value)
    if not text:
        return None
    chunks = text.split(":")
    if len(chunks) == 2:
        return f"{chunks[0].zfill(2)}:{chunks[1].zfill(2)}:00"
    if len(chunks) == 3:
        return f"{chunks[0].zfill(2)}:{chunks[1].zfill(2)}:{chunks[2].zfill(2)}"
    return None


def load_rows(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    items = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        item = {}
        for index, header in enumerate(headers):
            item[header] = row[index] if index < len(row) else None
        items.append(item)
    return items


def build_client_map(rows):
    clients = {}
    for row in rows:
        rut = normalize_rut(row.get("RUT"))
        if not rut:
            continue
        clients[rut] = {
            "email": normalize_email(row.get("Email")),
            "nombre": normalize_text(row.get("Nombres")),
            "apellido": normalize_text(row.get("Apellidos")),
            "telefono": normalize_phone(row.get("Teléfono")),
        }
    return clients


def build_reservations(rows, client_map):
    reservations = []
    for index, row in enumerate(rows, start=2):
        rut = normalize_rut(row.get("RUT"))
        client = client_map.get(rut, {})
        reservation = {
            "excel_row": index,
            "fecha_inicio": normalize_date(row.get("FECHA INICIO")),
            "fecha_finalizacion": normalize_date(row.get("FECHA FINAL")),
            "hora_inicio": normalize_time(row.get("HORA INICIO")),
            "hora_finalizacion": normalize_time(row.get("HORA FINAL")),
            "nombre": normalize_text(row.get("Nombre")) or client.get("nombre"),
            "apellido": normalize_text(row.get("Apellido")) or client.get("apellido"),
            "email": normalize_email(row.get("E-mail")) or client.get("email"),
            "telefono": normalize_phone(row.get("Teléfono")) or client.get("telefono"),
            "rut": rut,
        }
        reservations.append(reservation)
    return reservations


def validate_row(row):
    required = [
        "fecha_inicio",
        "fecha_finalizacion",
        "hora_inicio",
        "hora_finalizacion",
        "nombre",
        "apellido",
        "email",
        "telefono",
        "rut",
    ]
    missing = [field for field in required if not row.get(field)]
    if missing:
        return f"Faltan datos: {', '.join(missing)}"
    if row["fecha_inicio"] != row["fecha_finalizacion"]:
        return "La reserva debe iniciar y terminar el mismo día"
    return None


def reservation_block(row):
    fecha_inicio = sql_literal(row["fecha_inicio"])
    fecha_finalizacion = sql_literal(row["fecha_finalizacion"])
    hora_inicio = sql_literal(row["hora_inicio"])
    hora_finalizacion = sql_literal(row["hora_finalizacion"])
    nombre = sql_literal(row["nombre"])
    apellido = sql_literal(row["apellido"])
    rut = sql_literal(row["rut"])
    telefono = sql_literal(row["telefono"])
    email = sql_literal(row["email"])

    return f"""-- Fila Excel {row['excel_row']} | {row['nombre']} {row['apellido']} | {row['rut']}
SET @fechaInicio = {fecha_inicio};
SET @fechaFinalizacion = {fecha_finalizacion};
SET @horaInicio = {hora_inicio};
SET @horaFinalizacion = {hora_finalizacion};
SET @nombrePaciente = {nombre};
SET @apellidoPaciente = {apellido};
SET @rutPaciente = {rut};
SET @telefonoPaciente = {telefono};
SET @emailPaciente = {email};
SET @estadoReserva = 'reservada';

INSERT INTO pacienteDatos (
    nombre,
    apellido,
    rut,
    nacimiento,
    sexo,
    prevision_id,
    telefono,
    correo,
    direccion,
    pais,
    observaciones
)
SELECT
    @nombrePaciente,
    @apellidoPaciente,
    @rutPaciente,
    NULL,
    NULL,
    0,
    @telefonoPaciente,
    @emailPaciente,
    NULL,
    'chile',
    NULL
FROM DUAL
WHERE NOT EXISTS (
    SELECT 1
    FROM pacienteDatos
    WHERE rut = @rutPaciente
);

SET @id_paciente = (
    SELECT id_paciente
    FROM pacienteDatos
    WHERE rut = @rutPaciente
      AND estado_paciente <> 0
    LIMIT 1
);

SET @disponible = (
    SELECT CASE
        WHEN COUNT(*) = 0 THEN 1
        ELSE 0
    END
    FROM reservaPacientes
    WHERE NOT (
        TIMESTAMP(fechaFinalizacion, horaFinalizacion) <= TIMESTAMP(@fechaInicio, @horaInicio)
        OR TIMESTAMP(fechaInicio, horaInicio) >= TIMESTAMP(@fechaFinalizacion, @horaFinalizacion)
    )
);

INSERT INTO reservaPacientes (
    nombrePaciente,
    apellidoPaciente,
    rut,
    telefono,
    email,
    fechaInicio,
    horaInicio,
    fechaFinalizacion,
    horaFinalizacion,
    estadoReserva
)
SELECT
    @nombrePaciente,
    @apellidoPaciente,
    @rutPaciente,
    @telefonoPaciente,
    @emailPaciente,
    @fechaInicio,
    @horaInicio,
    @fechaFinalizacion,
    @horaFinalizacion,
    @estadoReserva
FROM DUAL
WHERE @disponible = 1;

SET @id_reserva_insertada = IF(ROW_COUNT() > 0, LAST_INSERT_ID(), NULL);

INSERT INTO fichaClinica (
    id_paciente,
    tipoAtencion,
    motivoConsulta,
    signosVitales,
    observaciones,
    anotacionConsulta,
    anamnesis,
    diagnostico,
    indicaciones,
    archivosAdjuntos,
    fechaConsulta,
    consentimientoFirmado
)
SELECT
    @id_paciente,
    'Agendamiento Automatico',
    NULL,
    NULL,
    NULL,
    CONCAT('Hora de Agendamiento: ', @horaInicio, ' - ', @horaFinalizacion),
    NULL,
    NULL,
    NULL,
    NULL,
    @fechaFinalizacion,
    NULL
FROM DUAL
WHERE @id_reserva_insertada IS NOT NULL
  AND @id_paciente IS NOT NULL;
"""


def main():
    clients = build_client_map(load_rows(CLIENTES_XLSX))
    reservations = build_reservations(load_rows(RESERVAS_XLSX), clients)

    valid_rows = []
    invalid_rows = []

    for row in reservations:
        error = validate_row(row)
        if error:
            invalid_rows.append({"excel_row": row["excel_row"], "rut": row["rut"], "error": error})
        else:
            valid_rows.append(row)

    blocks = [reservation_block(row) for row in valid_rows]

    sql = f"""-- Script generado automaticamente desde los Excel de Desktop
-- Fecha de generacion: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
-- Fuente reservas: {RESERVAS_XLSX.name}
-- Fuente clientes: {CLIENTES_XLSX.name}
-- Filas validas: {len(valid_rows)}
-- Filas invalidas: {len(invalid_rows)}

START TRANSACTION;

{chr(10).join(blocks)}
COMMIT;
"""

    OUTPUT_SQL.write_text(sql, encoding="utf-8")
    OUTPUT_JSON.write_text(json.dumps({
        "valid_rows": len(valid_rows),
        "invalid_rows": invalid_rows,
    }, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"SQL generado en: {OUTPUT_SQL}")
    print(f"Preview generado en: {OUTPUT_JSON}")
    print(f"Filas válidas: {len(valid_rows)}")
    print(f"Filas inválidas: {len(invalid_rows)}")


if __name__ == "__main__":
    main()
